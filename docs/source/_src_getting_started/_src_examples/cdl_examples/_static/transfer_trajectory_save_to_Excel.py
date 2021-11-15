import os
import pandas as pd
import ast
import xlsxwriter
from openpyxl import load_workbook
from transfer_trajectory_inputs import transfer_body_orders, leg_type
from TransferTrajectoryUtilities import get_transfer_body_order_abbreviation, get_user_input_for_saving
from tudatpy.kernel.astro import trajectory_design

def main():

    print()
    print('This script will save trajectory analysis results to Excel')
    print()

    current_dir = os.getcwd()

    # Set the destination directory for the Excel file
    excel_path = current_dir + '/Transfer_trajectory.xlsx'

    # Create the Excel file if it does not exist yet
    if not os.path.isfile(excel_path):
        os.mknod(excel_path)

    # Get user input on which transfer trajectory case to save
    dir_to_save = get_user_input_for_saving(transfer_body_orders,
                                            current_dir,
                                            leg_type)

    # Read trajectory information (delta V and TOF)
    f = open(dir_to_save + '/trajectory_info.txt')
    lines = f.readlines()
    f.close()
    tof = float(lines[1].split(':')[1].strip('/n'))
    dv = float(lines[2].split(':')[1].strip('/n'))
    dv_per_leg = lines[3].split(':')[1].strip('/n')
    dv_per_leg = ast.literal_eval(dv_per_leg[1:])
    dv_per_node = lines[4].split(':')[1].strip('/n')
    dv_per_node = ast.literal_eval(dv_per_node[1:])

    # Save trajectory information to Excel
    wb = xlsxwriter.Workbook(excel_path)
    ws = wb.add_worksheet('Trajectory info')
    ws.write(0, 0, 'TOF [days]')
    ws.write(0, 1, tof)
    ws.write(1, 0, 'Delta V [m/s]')
    ws.write(1, 1, dv)
    ws.write(2, 0, 'Delta V per leg [m/s]')
    for i in range(len(dv_per_leg)):
        ws.write(2, i+1, dv_per_leg[i])
    ws.write(3, 0, 'Delta V per node [m/s]')
    for i in range(len(dv_per_node)):
        ws.write(3, i + 1, dv_per_node[i])
    wb.close()

    # Set pandas excel writer to save analysis results
    book = load_workbook(excel_path)
    writer = pd.ExcelWriter(excel_path, engine='openpyxl')
    writer.book = book

    # Read analysis results
    comms_time = pd.read_csv(dir_to_save + '/trajectory_analysis_data/comms_time.txt', sep=' ')
    link_budget = pd.read_csv(dir_to_save + '/trajectory_analysis_data/link_budget.txt', sep=' ')
    solar_flux = pd.read_csv(dir_to_save + '/trajectory_analysis_data/solar_flux.txt', sep=' ')
    state_wrt_earth = pd.read_csv(dir_to_save + '/trajectory_analysis_data/state_wrt_earth.txt', sep=' ')
    state_wrt_sun = pd.read_csv(dir_to_save + '/trajectory_analysis_data/state_wrt_sun.txt', sep=' ')

    # Save analysis results to Excel
    comms_time.to_excel(writer, sheet_name='Communication time', columns=['#', 'time[s]'], header=['Time [s]', 'Communication time per day [s]'], index=False, freeze_panes=(1,0))
    link_budget.to_excel(writer, sheet_name='Link budget', columns=['#', 'time[s]'], header=['Time [s]', 'Link budget [W]'], index=False, freeze_panes=(1,0))
    solar_flux.to_excel(writer, sheet_name='Solar Flux', columns=['#', 'time[s]'], header=['Time [s]', 'Solar Flux [W/m2]'], index=False, freeze_panes=(1,0))
    state_wrt_earth.to_excel(writer, sheet_name='State wrt Earth', columns=['#', 'time[s]', '\t', 'x[m]', '\t.1', 'y[m]', '\t.2'], header=['Time [s]', 'x [m]', 'y [m]', 'z [m]', 'vx [m/s]', 'vy [m/s]', 'vz [m/s]'], index=False, freeze_panes=(1,0))
    state_wrt_sun.to_excel(writer, sheet_name='State wrt Sun', columns=['#', 'time[s]', '\t', 'x[m]', '\t.1', 'y[m]', '\t.2'], header=['Time [s]', 'x [m]', 'y [m]', 'z [m]', 'vx [m/s]', 'vy [m/s]', 'vz [m/s]'], index=False, freeze_panes=(1,0))

    # Save the file
    writer.save()

    return 0

if __name__ == "__main__":    main()

